import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import streamlit as st

# Streamlit App
st.title("Repetitive Value Finder Across Files")

# File Uploader
uploaded_files = st.file_uploader("Upload CSV files", accept_multiple_files=True, type=["csv"])

# Column Name Input
column_name = st.text_input("Enter the column name you want to compare across the files:")

if uploaded_files and column_name:
    column_name = column_name.strip()  # Clean up any extra spaces
    data = {}  # Dictionary to store column data

    for uploaded_file in uploaded_files:
        try:
            # Read the uploaded CSV file
            df = pd.read_csv(uploaded_file)
            if column_name in df.columns:  # Ensure the specified column exists
                file_key = uploaded_file.name  # Use file name as the key
                data[file_key] = df[column_name].dropna()  # Drop NaN values
            else:
                st.warning(f"'{column_name}' column not found in {uploaded_file.name}")
        except Exception as e:
            st.error(f"Error processing file {uploaded_file.name}: {e}")

    # Process the data if there are valid files with the column
    if data:
        # Combine all data into a single Series with file names as the index
        combined_data = pd.concat(data, names=["File"]).reset_index()

        # Identify duplicate values across files
        duplicate_values = combined_data[combined_data.duplicated(subset=[column_name], keep=False)]

        # Group duplicates to show which files they appear in
        result = duplicate_values.groupby(column_name)["File"].apply(lambda x: ", ".join(set(x))).reset_index()

        # Save the result to an Excel file with yellow highlighting for repetitive entries
        output_file = "repetitive_values.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Repetitive Values"

        # Add header row
        header = [column_name, "Files"]
        for col_num, value in enumerate(header, 1):
            sheet.cell(row=1, column=col_num, value=value)

        # Add data rows and highlight repetitive entries
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row_num, (value, files) in enumerate(result.values, start=2):
            sheet.cell(row=row_num, column=1, value=value)  # Repetitive Value from the specified column
            sheet.cell(row=row_num, column=2, value=files)  # Files containing the repetitive value
            # Highlight if the value appears in more than one file
            if "," in files:
                sheet.cell(row=row_num, column=1).fill = yellow_fill
                sheet.cell(row=row_num, column=2).fill = yellow_fill

        # Save the Excel file
        workbook.save(output_file)

        # Provide download link for the result file
        with open(output_file, "rb") as f:
            st.download_button(
                label="Download Repetitive Values Excel File",
                data=f,
                file_name=output_file,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    else:
        st.warning(f"No valid '{column_name}' data found in the uploaded files.")
